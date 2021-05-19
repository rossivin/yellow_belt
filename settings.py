import openpyxl as xl

class Settings():

	def __init__(self):

		#INITIALIZING EXCEL WORKSHEETS
		self.data_workbook = "data_results.xlsx"
		self.data_sheet = "Data"
		self.brand_sheet = "Brand Matrix"
		self.matrix_sheet = "Compatibility Matrix"

		#FILTER SETTINGS
		filter1 = {"volume": 40, "pre_volume": 12, "post_volume": 30}
		filter2 = {"volume": 40, "pre_volume": 12, "post_volume": 30}
		self.filter_specs = [filter1, filter2]

		#INTERFACE SETTINGS
		switch1 = {0: 70, 1: 70, 2: 50, 3: 0, 4: 0}
		switch2 = {0: 70, 1: 70, 2: 50, 3: 0, 4: 0}
		self.interface_specs = [switch1, switch2]

		#BRAND INFORMATION
		self.brand_dictionary = self.getBrandDictionary()
		self.compatibility_matrix = self.getCompatibilityMatrix()

	def getBrandDictionary(self):

		brand_dictionary = {}

		brands_workbook = xl.load_workbook(filename = self.data_workbook)
		sheet = brands_workbook[self.brand_sheet]

		for r in range(2,37):
			data_dictionary = {}
			brand_name = str(sheet.cell(row = r, column = 5).value)
			data_dictionary["alcohol"] = sheet.cell(row = r, column = 7).value
			data_dictionary["cost"] = sheet.cell(row = r, column = 8).value
			brand_dictionary[brand_name] = data_dictionary

		return brand_dictionary

	def getCompatibilityMatrix(self):
		
		matrix = {}
		brands_workbook = xl.load_workbook(filename = self.data_workbook)
		sheet = brands_workbook[self.matrix_sheet]

		for r in range(2, sheet.max_row + 1):
			brand1_dictionary = {}
			name1 = sheet.cell(row = r, column = 1).value
			for c in range(2, sheet.max_row + 1):
				name2 = sheet.cell(row = 1, column = c).value
				compatibility_value = sheet.cell(row = r, column = c).value
				brand1_dictionary[name2] = compatibility_value
			matrix[name1] = brand1_dictionary	
		return matrix


