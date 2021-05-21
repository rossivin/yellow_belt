from brand import Brand
from filter_vessel import FilterVessel
from filter_interface import FilterInterface
import openpyxl as xl

"""INPUTS"""
#brand1
brand1_name = "Budweiser"
brand1_cost = 22.50
brand1_abv = 5.0
#brand2
brand2_name = "Bud Light"
brand2_cost = 30.00
brand2_abv = 6.0
#filter
filter_volume = 40.0
pre_filter_volume = 10.0
post_filter_volume = 20.0
alcoholyzer_location = 15.0
#parameters
switch_volume = 25
volume_interval = 1

"""INITIALIZING"""
brand1 = Brand(brand1_name, brand1_cost, brand1_abv)
brand2 = Brand(brand2_name, brand2_cost, brand2_abv)
vessel = FilterVessel(filter_volume, pre_filter_volume, post_filter_volume)

def simulate_instant(switch_volume, vessel, brand1, brand2):
	interface_sim = FilterInterface(vessel, brand1, brand2)
	print("Filter Size: %.2f hL" % filter_volume)
	print("Pre-Filter Volume: %.2f hL" % pre_filter_volume)
	print("Post-Filter Volume: %.2f hL" % post_filter_volume)

	print("\nTotal Volume of Brand 1 ({}) in Brand 2 ({}): {:.2f} hL".format(brand1.name, brand2.name, interface_sim.brandVolumeLeft(1, switch_volume)))
	print("Total Volume of Brand 2 ({}) in Brand 1 ({}): {:.2f} hL".format(brand2.name, brand1.name, interface_sim.totalBrandVolumeOut(2, switch_volume)))
	print("Financial Impact of Switch: CAD {:.2f}".format(interface_sim.financialImpactOfSwitch(switch_volume)))
	print("Alcoholyzer Value: {:.2f}%".format(interface_sim.alcoholyzerValue(switch_volume, alcoholyzer_location)))

def simulate_excel(vessel, brand1, brand2):
	interface_sim = FilterInterface(vessel, brand1, brand2)
	
	wb = xl.Workbook()
	sheet = wb.active
	excel_headings(sheet)

	current_volume = 0
	r = 4 #Row

	while interface_sim.outletComposition(1, current_volume) > 0.01:
		sheet.cell(row = r, column = 1).value = current_volume
		sheet.cell(row = r, column = 2).value = interface_sim.brandVolumeLeft(1, current_volume)
		sheet.cell(row = r, column = 3).value = interface_sim.totalBrandVolumeOut(2, current_volume)
		sheet.cell(row = r, column = 4).value = interface_sim.financialImpactOfSwitch(current_volume)
		sheet.cell(row = r, column = 5).value = interface_sim.alcoholyzerValue(current_volume, alcoholyzer_location)

		current_volume += volume_interval
		r += 1

	wb.save("simulation_results.xlsx")

def excel_headings(sheet):
	sheet.cell(row = 1, column = 1).value = "Filter Vol: {}hL, Pre-Filter Vol: {}hL, Post-Filter Vol: {}hL, Brand 1: {}, ABV: {}%, Cost: {}$/hL, Brand 2: {}, ABV: {}%, Cost: {}$/hL".format(filter_volume,pre_filter_volume,post_filter_volume,brand1_name, brand1_abv, brand1_cost, brand2_name, brand2_abv, brand2_cost)
	sheet.cell(row = 3, column = 1).value = "Volume Pushed"
	sheet.cell(row = 3, column = 2).value = "Volume of Brand 1 Left"
	sheet.cell(row = 3, column = 3).value = "Volume of Brand 2 Out"
	sheet.cell(row = 3, column = 4).value = "Financial Impact of Switch"
	sheet.cell(row = 3, column = 5).value = "Alcoholyzer Reading"

simulate_instant(switch_volume, vessel, brand1, brand2)

simulate_excel(vessel, brand1, brand2)









